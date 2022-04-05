from django.http import HttpResponse
from solo.models import SingletonModel
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.utils.timezone import now, datetime


class OrdersInTimeStatistic(SingletonModel):

    class Meta:
        verbose_name = '03 - Статистика по общему количеству заказов'

    @classmethod
    def generate_doc(cls, datetime_from, datetime_to,
                     wholesalers, dropshippers, retailers):

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Статистика по общему количеству заказов ' \
                          '(период, фирма, логины, категории пользователей)'
        worksheet['A1'] = 'Дата от: '
        worksheet['A2'] = datetime_from
        worksheet['A3'] = 'Дата до:'
        worksheet['A4'] = datetime_to
        worksheet['B1'] = 'Оптовики'
        worksheet['B2'] = 'Пользователи'
        worksheet['C2'] = 'Их бренды'
        worksheet['D2'] = 'количество заказов этого бренда'
        worksheet['E2'] = 'Всего заказов'

        row = 3
        for dicts in wholesalers:
            # worksheet[f'C{row}'] = ''
            brands = {}
            for items_or_lists in dicts.values():
                if type(items_or_lists) is list:
                    for current_dict in items_or_lists:
                        brand_name = ''
                        for keys, values in current_dict.items():
                            if type(values) is not int:
                                brand_name += values
                            else:
                                brands[brand_name] = values
                                brand_name = ''
                else:
                    if type(items_or_lists) is int or len(items_or_lists) == 1:
                        worksheet.cell(row=row, column=5, value=str(items_or_lists))
                        for keys, values in brands.items():
                            worksheet.cell(row=row, column=3, value=str(keys))
                            worksheet.cell(row=row, column=4, value=str(values))
                            row += 1
                        row += 1
                    else:
                        worksheet.cell(row=row, column=2, value=str(items_or_lists))
        worksheet['F1'] = 'Дропшипперы'
        worksheet['F2'] = 'Пользователи'
        worksheet['G2'] = 'Их бренды'
        worksheet['H2'] = 'количество заказов этого бренда'
        worksheet['I2'] = 'Всего заказов'

        row = 3
        for dicts in dropshippers:
            # worksheet[f'C{row}'] = ''
            brands = {}
            for items_or_lists in dicts.values():
                if type(items_or_lists) is list:
                    for current_dict in items_or_lists:
                        brand_name = ''
                        for keys, values in current_dict.items():
                            if type(values) is not int:
                                brand_name += values
                            else:
                                brands[brand_name] = values
                                brand_name = ''
                else:
                    if type(items_or_lists) is int or len(items_or_lists) == 1:
                        worksheet.cell(row=row, column=9, value=str(items_or_lists))
                        for keys, values in brands.items():
                            worksheet.cell(row=row, column=7, value=str(keys))
                            worksheet.cell(row=row, column=8, value=str(values))
                            row += 1
                        row += 1
                    else:
                        worksheet.cell(row=row, column=6, value=str(items_or_lists))
        worksheet['J1'] = 'Розница'
        worksheet['J2'] = 'Пользователи'
        worksheet['K2'] = 'Их бренды'
        worksheet['L2'] = 'количество заказов этого бренда'
        worksheet['M2'] = 'Всего заказов'

        row = 3
        for dicts in retailers:
            brands = {}
            for items_or_lists in dicts.values():
                if type(items_or_lists) is list:
                    for current_dict in items_or_lists:
                        brand_name = ''
                        for keys, values in current_dict.items():
                            if type(values) is not int:
                                brand_name += values
                            else:
                                brands[brand_name] = values
                                brand_name = ''
                else:
                    if type(items_or_lists) is int or len(items_or_lists) == 1:
                        worksheet.cell(row=row, column=13, value=str(items_or_lists))
                        for keys, values in brands.items():
                            worksheet.cell(row=row, column=11, value=str(keys))
                            worksheet.cell(row=row, column=12, value=str(values))
                            row += 1
                        row += 1
                    else:
                        worksheet.cell(row=row, column=10, value=str(items_or_lists))

        current_datetime = now()
        workbook.save(f'public/media/Статистика {current_datetime}.xlsx')
        current_datetime = str(current_datetime)
        current_datetime = current_datetime[0:16]

        response = HttpResponse(content=save_virtual_workbook(workbook))
        # mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response['Content-Disposition'] = f'attachment; filename={current_datetime}.xlsx'
        return response
